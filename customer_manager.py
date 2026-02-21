import sqlite3
import os
from openpyxl import Workbook, load_workbook

DB_NAME = "jewellery.db"
EXCEL_FILE = "customers_backup.xlsx"


# -------------------------------
# Database Connection
# -------------------------------
def connect_db():
    return sqlite3.connect(DB_NAME)


# -------------------------------
# Create Customer Table
# -------------------------------
def create_customer_table():
    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            mobile TEXT UNIQUE NOT NULL,
            city TEXT,
            address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.commit()
    conn.close()


# -------------------------------
# Create Excel File If Not Exists
# -------------------------------
def create_excel_if_not_exists():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws.append(["Name", "Mobile", "City", "Address"])
        wb.save(EXCEL_FILE)


# -------------------------------
# Backup / Update Customer in Excel
# -------------------------------
def backup_customer_to_excel(name, mobile, city, address):
    create_excel_if_not_exists()

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    mobile = str(mobile).strip()

    for row in ws.iter_rows(min_row=2):
        if str(row[1].value).strip() == mobile:
            row[0].value = name
            row[2].value = city
            row[3].value = address
            wb.save(EXCEL_FILE)
            wb.close()
            return

    ws.append([name, mobile, city, address])
    wb.save(EXCEL_FILE)
    wb.close()


# -------------------------------
# Add or Update Customer (Main)
# -------------------------------
def add_customer(name, mobile, city, address):
    mobile = str(mobile).strip()

    conn = connect_db()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            INSERT INTO customers (name, mobile, city, address)
            VALUES (?, ?, ?, ?)
        """, (name, mobile, city, address))

        print("Customer Inserted")

    except sqlite3.IntegrityError:
        cursor.execute("""
            UPDATE customers
            SET name = ?, city = ?, address = ?
            WHERE mobile = ?
        """, (name, city, address, mobile))

        if cursor.rowcount > 0:
            print("Customer Updated")
        else:
            print("Update Failed - Mobile Not Found")

    conn.commit()
    conn.close()

    # Backup to Excel
    backup_customer_to_excel(name, mobile, city, address)


# -------------------------------
# Check if Customer Exists
# -------------------------------
def customer_exists(mobile):
    mobile = str(mobile).strip()

    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM customers WHERE mobile = ?", (mobile,))
    result = cursor.fetchone()

    conn.close()
    return result is not None


# -------------------------------
# Get Full Customer Record
# -------------------------------
def get_customer(mobile):
    mobile = str(mobile).strip()

    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM customers WHERE mobile = ?", (mobile,))
    result = cursor.fetchone()

    conn.close()
    return result


# -------------------------------
# Get Customer by Mobile (Formatted)
# -------------------------------
def get_customer_by_mobile(mobile):
    mobile = str(mobile).strip()

    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT name, city, address 
        FROM customers 
        WHERE mobile = ?
    """, (mobile,))

    result = cursor.fetchone()
    conn.close()

    if result:
        return {
            "name": result[0],
            "city": result[1],
            "address": result[2]
        }

    return None


# -------------------------------
# Search Customers (Autocomplete)
# -------------------------------
def search_customers(query):
    query = str(query).strip()

    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT name, mobile 
        FROM customers
        WHERE mobile LIKE ? OR name LIKE ?
        ORDER BY created_at DESC
        LIMIT 5
    """, (f"%{query}%", f"%{query}%"))

    results = cursor.fetchall()
    conn.close()

    return [
        {"name": row[0], "mobile": row[1]}
        for row in results
    ]
def get_all_customers():
    conn = get_db_connection()
    customers = conn.execute("SELECT * FROM customers ORDER BY name").fetchall()
    conn.close()
    return customers